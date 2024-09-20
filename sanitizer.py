from typing import Dict, List, Tuple
import datetime
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def is_english(text):
    # Check if the text contains only English alphabet characters and spaces
    return bool(re.match(r"^[A-Za-z\s]*$", text))


NAME_KEY = "Name"
TIMESTAMP_KEY = "חותמת זמן"


def read_results(filename: str) -> List[Dict[str, str]]:
    with open(filename, newline="", encoding="utf-8-sig") as data:
        reader = csv.DictReader(data)
        return [row for row in reader]


def sanitize_cell(cell: str, column: str) -> Tuple[str, bool]:
    if column == NAME_KEY:
        return sanitize_name(cell)
    if column == TIMESTAMP_KEY:
        return cell, True
    return sanitize_time(column, cell)


def sanitize_name(raw_name: str) -> Tuple[str, bool]:
    stripped_string = raw_name.strip()

    if not is_english(stripped_string):
        return stripped_string, False

    name_tokens = [word.capitalize() for word in stripped_string.split()]
    if len(name_tokens) < 2:
        return stripped_string, False
    return " ".join(name_tokens), True


def is_time_duration(raw_time: str) -> bool:
    if raw_time == "":
        return True

    try:
        if "." in raw_time:
            if ":" in raw_time:
                # Format is minutes:seconds.milliseconds
                minutes, seconds = raw_time.split(":")
                seconds = seconds.split(".")[0]  # Get only the seconds part
                milliseconds = seconds.split(".")[1] if "." in seconds else "00"
            else:
                # Format is minutes.seconds
                minutes, milliseconds = raw_time.split(".")
                seconds = "00"
        else:
            return False  # Must have either a dot or a colon

        # Convert minutes and seconds to integers
        minutes = int(minutes)
        seconds = int(seconds)
        milliseconds = int(milliseconds)

        # Check if minutes are valid (0-59) and seconds are valid (0-59)
        return 0 <= minutes < 60 and 0 <= seconds < 60 and 0 <= milliseconds < 100

    except (ValueError, IndexError):
        return False


def is_multiblind_result(input_str: str) -> bool:
    pattern = r"^(\d+)/(\d+) (\S+)$"

    match = re.match(pattern, input_str)
    if not match:
        return False

    solved = int(match.group(1))
    total = int(match.group(2))
    timestamp = match.group(3)

    if solved > total:
        return False

    return is_time_duration(timestamp)


def is_fmc_result(input_str: str) -> bool:
    return input_str.isdigit()


def sanitize_time(event: str, raw_time: str) -> Tuple[str, bool]:
    stripped_string = raw_time.strip()
    if stripped_string == "":
        return "", True
    if stripped_string.upper() == "DNF":
        return "DNF", True
    event_result_validators = {"Multiblind": is_multiblind_result, "FMC": is_fmc_result}
    result_parser = event_result_validators.get(event, is_time_duration)
    if not result_parser(stripped_string):
        return stripped_string, False
    return stripped_string, True


def write_sanitized(data: List[Dict[str, str]], filename: str):
    if not data:
        print("the provided list is empty")
        return
    headers = data[0].keys()
    with open(filename, mode="w", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)


def main():
    with open("./results.csv", newline="", encoding="utf-8-sig") as data:
        reader = csv.DictReader(data)
        rows = [row for row in reader]

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the header
    headers = list(rows[0].keys())
    ws.append(headers)

    # Define fill colors for different states
    fill_not_ok = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )  # Red for not ok

    for row in rows:
        for column, cell in row.items():
            sanitized, ok = sanitize_cell(cell, column)

    # Write the data and apply colors based on ok status
    for idx, row in enumerate(rows, start=2):  # start from row 2 since row 1 is header
        for column, cell in row.items():
            sanitized, ok = sanitize_cell(cell, column)
            cell_to_write = ws.cell(
                row=idx, column=list(row.keys()).index(column) + 1, value=sanitized
            )

            # Apply color based on ok variable
            if not ok:
                cell_to_write.fill = fill_not_ok  # Apply red fill if not ok

    # Save the workbook
    wb.save("output.xlsx")


if __name__ == "__main__":
    main()
