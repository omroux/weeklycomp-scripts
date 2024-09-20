from typing import Dict, List, Tuple
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def is_english(text):
    # Check if the text contains only English alphabet characters and spaces
    return bool(re.match(r"^[A-Za-z\s]*$", text))


NAME_KEY = "Name"
TIMESTAMP_KEY = "חותמת זמן"


def read_results(filename: str) -> List[Dict[str, str]]:
    with open(filename, newline="", encoding="utf-8-sig") as data:
        reader = csv.DictReader(data)
        return [row for row in reader]


def validate_cell(cell: str, column: str) -> Tuple[str, bool]:
    """
    Validate a cell based on its column, signifying its type (name, event etc)
    """
    if column == NAME_KEY:
        return validate_name(cell)
    if column == TIMESTAMP_KEY:
        return cell, True
    return validate_time(column, cell)


def validate_name(raw_name: str) -> Tuple[str, bool]:
    """
    Validate a competitor's name such that:
        - The name is in english
        - Every word's first letter is capitalized
        - The name is at least 2 words long
    """
    stripped_string = raw_name.strip()

    if not is_english(stripped_string):
        return stripped_string, False

    name_tokens = [word.capitalize() for word in stripped_string.split()]
    if len(name_tokens) < 2:
        return stripped_string, False
    return " ".join(name_tokens), True


def is_time_duration(raw_time: str) -> bool:
    """
    Validates if a given string represents a time duration in a specific format.

    The function accepts a time duration in two main formats:
        - "minutes:seconds.milliseconds"
        - "minutes.seconds"

    The following criteria must be met for the input to be considered valid:
        - Minutes must be an integer between 0 and 59 (inclusive)
        - Seconds must be an integer between 0 and 59 (inclusive)
        - Milliseconds must be an integer between 0 and 99 (inclusive)
    """
    if raw_time == "":
        return True

    try:
        if "." in raw_time:
            if ":" in raw_time:
                minutes, seconds = raw_time.split(":")
                seconds = seconds.split(".")[0]  # Get only the seconds part
                milliseconds = seconds.split(".")[1] if "." in seconds else "00"
            else:
                minutes, milliseconds = raw_time.split(".")
                seconds = "00"
        else:
            return False

        minutes = int(minutes)
        seconds = int(seconds)
        milliseconds = int(milliseconds)

        return 0 <= minutes < 60 and 0 <= seconds < 60 and 0 <= milliseconds < 100

    except (ValueError, IndexError):
        return False


def is_multiblind_result(input_str: str) -> bool:
    """
    Validate that the given input string represents a multiblind result.

    The function checks for this format:
        <solved_cubes>/<attempted_cubes> <duration>
    For example:
        1/2 12:34.56

    The duration is checked with the `is_time_duration` function.

    The function also checks that solved_cubes is lesser or equal to the attempted cubes.
    """
    pattern = r"^(\d+)/(\d+) (\S+)$"

    match = re.match(pattern, input_str)
    if not match:
        return False

    solved = int(match.group(1))
    total = int(match.group(2))
    timestamp = match.group(3)

    if solved > total:
        return False

    return is_time_duration(timestamp)


def is_fmc_result(input_str: str) -> bool:
    """
    Validate that the given input string represents an FMC result.

    An FMC result is just a string of digits.
    """
    return input_str.isdigit()


def validate_time(event: str, raw_time: str) -> Tuple[str, bool]:
    """
    Function to validate times, based on the given event.
    """
    stripped_string = raw_time.strip()
    if stripped_string == "":
        return "", True
    if stripped_string.upper() == "DNF":
        return "DNF", True
    event_result_validators = {"Multiblind": is_multiblind_result, "FMC": is_fmc_result}
    result_parser = event_result_validators.get(event, is_time_duration)
    if not result_parser(stripped_string):
        return stripped_string, False
    return stripped_string, True


def main():
    rows = read_results("./results.csv")

    wb = Workbook()
    ws = wb.active

    headers = list(rows[0].keys())
    ws.append(headers)

    error_fill_color = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )

    for idx, row in enumerate(rows, start=2):
        for column, cell in row.items():
            sanitized, ok = validate_cell(cell, column)
            cell_to_write = ws.cell(
                row=idx, column=list(row.keys()).index(column) + 1, value=sanitized
            )

            if not ok:
                cell_to_write.fill = error_fill_color

    wb.save("output.xlsx")


if __name__ == "__main__":
    main()
